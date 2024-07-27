# 导入必要的库
import numpy as np
from scipy.interpolate import CubicSpline, UnivariateSpline, PchipInterpolator
import re
from scipy.optimize import leastsq, minimize
from openpyxl import load_workbook
import os
import json
import pandas as pd
from scipy.signal import savgol_filter, argrelextrema

# 插值函数参数
file_path ='data//施工图-3Y-路线逐桩坐标表.xlsx'
method = 'pchip'
Univariate_s = 0.05
isDrawRadius = False

def interpolate(x, y):
    """
    对给定的 x 和 y 数据进行插值。

    参数:
    x (list or np.ndarray): 自变量数据。
    y (list or np.ndarray): 因变量数据。

    返回:
    function: 返回插值函数，可以用于计算任意 x 值对应的 y 值。
    """
    # 将 x 和 y 转换为 numpy 数组
    x = np.array(x)
    y = np.array(y)

    # 检查 x 是否是严格递减的，如果是则翻转 x 和 y
    if np.all(np.diff(x) < 0):
        x = x[::-1]
        y = y[::-1]

    # 对 x 和 y 进行排序
    sorted_indices = np.argsort(x)  # 获取排序索引
    x_sorted = x[sorted_indices]    # 根据索引对 x 进行排序
    y_sorted = y[sorted_indices]    # 根据索引对 y 进行排序

    # 根据选择的插值方法进行插值
    if method == 'cubic':
        # 使用 Cubic Spline 方法进行插值
        f = CubicSpline(x_sorted, y_sorted)
    elif method == 'univariate':
        # 使用 Univariate Spline 方法进行插值
        f = UnivariateSpline(x_sorted, y_sorted, s=Univariate_s)
    elif method == 'pchip':
        # 使用 PCHIP 方法进行插值
        f = PchipInterpolator(x_sorted, y_sorted)
    else:
        # 如果指定的方法无效，抛出错误
        raise ValueError("Invalid method specified")
    
    # 返回插值函数
    return f

def calc_R(xc, yc, x, y):
    """
    计算每个点到圆心 (xc, yc) 的距离。

    参数:
    xc (float): 圆心的 x 坐标。
    yc (float): 圆心的 y 坐标。
    x (numpy.ndarray): 所有点的 x 坐标。
    y (numpy.ndarray): 所有点的 y 坐标。

    返回:
    numpy.ndarray: 包含每个点到圆心距离的数组。
    """
    return np.sqrt((x - xc)**2 + (y - yc)**2)

def f_2(c, x, y):
    """
    计算每个点到圆心距离的偏差。

    参数:
    c (tuple): 包含圆心的坐标 (xc, yc)。
    x (numpy.ndarray): 所有点的 x 坐标。
    y (numpy.ndarray): 所有点的 y 坐标。

    返回:
    numpy.ndarray: 每个点到圆心距离的偏差。
    """
    Ri = calc_R(c[0], c[1], x, y)
    return Ri - Ri.mean()

def normalize(x, y):
    """
    对输入的点集进行归一化处理，将 x 和 y 值缩放到 [0, 1] 的范围内。

    参数:
    x (numpy.ndarray): 所有点的 x 坐标。
    y (numpy.ndarray): 所有点的 y 坐标。

    返回:
    tuple: (缩放后的 x 坐标数组, 缩放后的 y 坐标数组, x 的最小值, x 的最大值, y 的最小值, y 的最大值)
    """
    x_min = np.min(x)
    x_max = np.max(x)
    y_min = np.min(y)
    y_max = np.max(y)
    
    x_scaled = (x - x_min) / (x_max - x_min)
    y_scaled = (y - y_min) / (y_max - y_min)
  
    return x_scaled, y_scaled, x_min, x_max, y_min, y_max

def denormalize(x_scaled, y_scaled, x_min, x_max, y_min, y_max):
    """
    对归一化后的点集进行反归一化处理，将 x 和 y 值恢复到原始范围内。

    参数:
    x_scaled (numpy.ndarray): 缩放后的 x 坐标数组。
    y_scaled (numpy.ndarray): 缩放后的 y 坐标数组。
    x_min (float): 原始 x 的最小值。
    x_max (float): 原始 x 的最大值。
    y_min (float): 原始 y 的最小值。
    y_max (float): 原始 y 的最大值。

    返回:
    numpy.ndarray, numpy.ndarray: 恢复后的 x 坐标数组和 y 坐标数组。
    """
    x = x_scaled * (x_max - x_min) + x_min
    y = y_scaled * (y_max - y_min) + y_min
  
    return x, y

import numpy as np

def move_method(x, y):
    """
    对输入的点集进行平移处理，将 x 和 y 值移动到原点附近。

    参数:
    x (numpy.ndarray): 所有点的 x 坐标。
    y (numpy.ndarray): 所有点的 y 坐标。

    返回:
    numpy.ndarray, numpy.ndarray: 平移后的 x 坐标数组和 y 坐标数组。
    """
    x_moved = x - x[-1]
    y_moved = y - y[-1]
  
    return x_moved, y_moved


def zero_segments_calculator(x, f, zero_segments, A=10):
    """
    计算零导数段的样条曲线段并返回。

    参数：
    x (numpy.ndarray): 自变量数组。
    f (callable): 函数或插值对象，接受 x 数组作为参数并返回对应的因变量数组。
    zero_segments (list): 包含元组 (start_index, end_index) 的列表，表示零导数段在 x 中的起始和结束索引。
    A (float, optional): 控制样条曲线绘制的参数，默认为 10。

    返回：
    tuple: 包含两个列表 x_segments 和 y_segments，分别表示每个零导数段的 x 和 y 值数组。
    """
    x_segments = []
    y_segments = []

    for segment in zero_segments:
        start_index, end_index = segment
        x_segment = x[start_index:end_index+1]
        y_segment = f(x_segment)
        x_segments.append(x_segment)
        y_segments.append(y_segment)

    return x_segments, y_segments

def non_zero_segments_calculator(x, f, combined_segments, A, B, x_min, x_max, y_min, y_max, type="基本对称", type1="无缓和曲线"):
    """
    计算非零导数段的样条曲线，并进行圆拟合。

    参数：
    x (numpy.ndarray): 自变量数组。
    f (callable): 函数或插值对象，接受 x 数组作为参数并返回对应的因变量数组。
    non_zero_segments (list): 包含元组 (start_index, end_index) 的列表，表示非零导数段在 x 中的起始和结束索引。
    A (list): 控制每个段样条曲线的参数列表。
    B (list or None): 可选参数，用于指定每个段的圆拟合中心索引，如果为 None，则自动选择最大曲率处作为曲率中心。

    返回：
    tuple: 包含四个列表 x_fits, y_fits, errors, centers，分别表示每个非零导数段的圆拟合 x 值数组、y 值数组、拟合误差和圆心索引。
    """
    x_fits = []
    y_fits = []
    errors = []
    centers = []
    tolerances = []

    # 取出所有段平曲线的索引数据
    non_zero_segments = [segment for segment, label in combined_segments if label == 'non_zero']

    for i, segment in enumerate(non_zero_segments):
        start_index, end_index = segment
        x_segment = x[start_index:end_index+1]
        y_segment = f(x_segment)

        # 计算曲率半径
        curvature = calculate_curvature(f, x_segment)

        # 找到曲率变化不大的点集
        avg_curvature = np.mean(curvature)
        tolerance = 0.3 * avg_curvature  # 10%的曲率容差，可根据需要调整
        consistent_curvature_indices = np.where(np.abs(curvature - avg_curvature) < tolerance)[0]

        if type1 == "无缓和曲线":
            consistent_curvature_indices = np.arange(len(x_segment))  # 生成从0到len(x_segment)-1的索引

        if len(consistent_curvature_indices) < 3:
            print('曲率变化不大的点集不足')
            continue

        # 确保点集的连续性
        continuous_segments = []
        current_segment = [consistent_curvature_indices[0]]
        for idx in consistent_curvature_indices[1:]:
            if idx == current_segment[-1] + 1:
                current_segment.append(idx)
            else:
                if len(current_segment) >= 3:  # 确保连续段至少包含3个点
                    continuous_segments.append(current_segment)
                current_segment = [idx]
        if len(current_segment) >= 3:
            continuous_segments.append(current_segment)

        if type == "基本对称":
            center_idx = (end_index - start_index) // 2
        elif type == "非基本对称":
            # 计算拟合范围内的中心点
            if B is None:
                center_idx = consistent_curvature_indices[len(consistent_curvature_indices) // 2]
            else:
                center_idx = B[i]


        # 查找 segment 在 combined_segments 中的索引
        combined_index = next(index for index, (seg, label) in enumerate(combined_segments) if seg == segment)

        # 判断当前 segment 是否前后都是 non_zero 类型的 segment
        if combined_index != -1:
            if combined_index > 0 and combined_index < len(combined_segments) - 1:
                previous_segment = combined_segments[combined_index - 1]
                next_segment = combined_segments[combined_index + 1]

                if previous_segment[1] == 'zero' and next_segment[1] == 'zero':
                    #起点直线的俩个点（pt2为计算点）
                    previous_pt1_x = x[previous_segment[0][1] - 1]
                    previous_pt1_x, previous_pt1_y = denormalize(previous_pt1_x, f(previous_pt1_x), x_min, x_max, y_min, y_max) 
                    previous_pt2_x = x[previous_segment[0][1]]
                    previous_pt2_x, previous_pt2_y = denormalize(previous_pt2_x, f(previous_pt2_x), x_min, x_max, y_min, y_max) 

                    # 计算切线的切向方向（弧度）(pt1指向pt2)
                    tangent_angle1 = np.arctan2(previous_pt2_y - previous_pt1_y, previous_pt2_x - previous_pt1_x)

                    #圆开头俩个点（pt1为计算点）
                    circle_start1 = max(0, center_idx - (len(consistent_curvature_indices) if A is None else A[i]))
                    circle_start2 = max(0, center_idx - (len(consistent_curvature_indices) if A is None else A[i])) + 1
                    circle_pt1_x, circle_pt1_y = denormalize(x_segment[circle_start1], y_segment[circle_start1], x_min, x_max, y_min, y_max) 
                    circle_pt2_x, circle_pt2_y = denormalize(x_segment[circle_start2], y_segment[circle_start2], x_min, x_max, y_min, y_max) 

                    #计算回旋线长度L
                    x1, y1 = denormalize(x_segment[0:circle_start1], y_segment[0:circle_start1], x_min, x_max, y_min, y_max) 
                    points1 = np.column_stack((x1, y1))
                    L1 = calculate_length(points1)

                    # 计算切线的切向方向（弧度）(pt1指向pt2)
                    tangent_angle2 = np.arctan2(circle_pt2_y - circle_pt1_y, circle_pt2_x - circle_pt1_x)

                    #A1 = clothoid_A(previous_pt2_x, previous_pt2_y, tangent_angle1, circle_pt1_x, circle_pt1_y, tangent_angle2)

                    #圆结束俩个点（pt4为计算点）
                    circle_end1 = min(len(x_segment), center_idx + (len(consistent_curvature_indices) if A is None else A[i])) - 1
                    circle_end2 = min(len(x_segment), center_idx + (len(consistent_curvature_indices) if A is None else A[i]))
                    circle_pt3_x, circle_pt3_y = denormalize(x_segment[circle_end1-1], y_segment[circle_end1-1], x_min, x_max, y_min, y_max) 
                    circle_pt4_x, circle_pt4_y = denormalize(x_segment[circle_end2-1], y_segment[circle_end2-1], x_min, x_max, y_min, y_max) 

                    # 计算切线的切向方向（弧度）(pt3指向pt4)
                    tangent_angle3 = np.arctan2(circle_pt4_y - circle_pt3_y, circle_pt4_x - circle_pt3_x)

                    #终点直线的俩个点（pt1为计算点）
                    next_pt1_x = x[next_segment[0][0]]
                    next_pt1_x, next_pt1_y = denormalize(next_pt1_x, f(next_pt1_x), x_min, x_max, y_min, y_max) 
                    next_pt2_x = x[next_segment[0][0] + 1]
                    next_pt2_x, next_pt2_y = denormalize(next_pt2_x, f(next_pt2_x), x_min, x_max, y_min, y_max) 
                    # 计算切线的切向方向（弧度）(pt1指向pt2)
                    tangent_angle4 = np.arctan2(next_pt2_y - next_pt1_y, next_pt2_x - next_pt1_x)

                    
                    #计算回旋线长度L
                    x2, y2 = denormalize(x_segment[circle_end2:-1], y_segment[circle_end2:-1], x_min, x_max, y_min, y_max) 
                    points2 = np.column_stack((x2, y2))
                    L2 = calculate_length(points2)

                    #A2 = clothoid_A(circle_pt4_x, circle_pt4_y, tangent_angle3, next_pt1_x, next_pt1_y, tangent_angle4)

                    pt1_x = x[previous_segment[0][0]]
                    pt1_x, pt1_y = denormalize(pt1_x, f(pt1_x), x_min, x_max, y_min, y_max) 

                    # 进行圆拟合
                    if A is None:
                        x_fit_circle, y_fit_circle, circle_error, R, xc, yc = circle_fit(center_idx, x_segment, y_segment, len(consistent_curvature_indices), x_min, x_max, y_min, y_max)
                        tolerances.append(len(consistent_curvature_indices))
                    else:
                        x_fit_circle, y_fit_circle, circle_error, R, xc, yc = circle_fit(center_idx, x_segment, y_segment, A[i], x_min, x_max, y_min, y_max)
                        tolerances = A

                    if x_fit_circle is not None and y_fit_circle is not None:
                        x_fits.append(x_fit_circle)
                        y_fits.append(y_fit_circle)
                        errors.append(circle_error)
                        centers.append(center_idx)
                    A1 = calculate_A(L1, R)
                    A2 = calculate_A(L2, R)
                    # print(f"第{i + 1}段直线起点坐标", pt1_x, pt1_y)
                    # print(f"第{i + 1}段直线终点坐标", previous_pt2_x, previous_pt2_y)
                    # print(f"第{i + 1}段平曲线第一部分回旋线值：(", "A值:", A1, "起点:", previous_pt2_x, previous_pt2_y, "终点:", circle_pt1_x, circle_pt1_y)
                    # print(f"第{i + 1}段平曲线第二部分回旋线值：(", "A值:", A2, "起点:", circle_pt4_x, circle_pt4_y, "终点:", next_pt1_x, next_pt1_y)
                    # print(f"第{i + 1}段平曲线的圆曲线值:")
                    # print("半径:", R, "圆心x:", xc, "圆心y:", yc, "\n\n")
                    # Store the data in a dictionary
                    data = {
                        "line_segment": {
                            "start_point": {"x": pt1_x, "y": pt1_y},
                            "end_point": {"x": previous_pt2_x, "y": previous_pt2_y}
                        },
                        "curve_segment_1": {
                            "A_value": A1,
                            "start_point": {"x": previous_pt2_x, "y": previous_pt2_y},
                            "end_point": {"x": circle_pt1_x, "y": circle_pt1_y}
                        },
                        "curve_segment_2": {
                            "A_value": A2,
                            "start_point": {"x": circle_pt4_x, "y": circle_pt4_y},
                            "end_point": {"x": next_pt1_x, "y": next_pt1_y}
                        },
                        "circle_curve": {
                            "radius": R,
                            "center": {"x": xc, "y": yc}
                        }
                    }
                    # Convert the dictionary to a JSON string
                    json_output = json.dumps(data, ensure_ascii=False, indent=4)

                    # Print the JSON string
                    #print(json_output)

    return x_fits, y_fits, tolerances, centers

def circle_fit(curvature_center, x, y, A, x_min, x_max, y_min, y_max):
    """
    对给定的数据点进行圆拟合。

    参数：
    curvature_center (int): 曲率中心的索引。
    x (numpy.ndarray): x 坐标数组。
    y (numpy.ndarray): y 坐标数组。
    A (int): 拟合范围长度。

    返回：
    tuple: 包含 x_fit_circle, y_fit_circle, fitting_error 的元组，分别表示拟合的圆的 x 坐标数组、y 坐标数组和拟合误差。

    注意：
    - 如果拟合点数量不足，将打印警告信息并返回 None。
    """
    # 提取拟合范围内的数据点
    x, y = denormalize(x, y, x_min, x_max, y_min, y_max) 
    fit_start = max(0, curvature_center - A)
    fit_end = min(len(x), curvature_center + A)
    x_fit = x[fit_start:fit_end]
    y_fit = y[fit_start:fit_end]

    # 检查拟合点的数量
    if len(x_fit) < 3:  # 至少需要三个点来拟合圆
        print('拟合点数量不足')
        return None, None, None

    # 初步估计圆心位置
    center_estimate = np.mean(x_fit), np.mean(y_fit)
    
    # 使用最小二乘法进行圆拟合
    center, ier = leastsq(f_2, center_estimate, args=(x_fit, y_fit))
    xc, yc = center
    
    # 计算拟合半径
    Ri = calc_R(xc, yc, x_fit, y_fit)
    R = Ri.mean()

    # real_x, real_y = denormalize(xc, yc, x_min, x_max, y_min, y_max) 

    # real_x_fit, real_y_fit = denormalize(x_fit, y_fit, x_min, x_max, y_min, y_max) 

    # Ri_real = calc_R(real_x, real_y, real_x_fit, real_y_fit)
    # R_real = Ri_real.mean()

    # 计算拟合误差
    fitting_error = np.sum((Ri - R)**2)
    
    # 根据拟合结果生成圆的坐标点
    theta_start = np.arctan2(y_fit[0] - yc, x_fit[0] - xc)
    theta_end = np.arctan2(y_fit[-1] - yc, x_fit[-1] - xc)
    theta_fit = np.linspace(theta_start, theta_end, 100)
    x_fit_circle = xc + R * np.cos(theta_fit)
    y_fit_circle = yc + R * np.sin(theta_fit)

    #real_x_fit_circle, real_y_fit_circle = denormalize(x_fit_circle, y_fit_circle, x_min, x_max, y_min, y_max) 
    # print("圆数据:")
    # for x, y in zip(x_fit_circle, x_fit_circle):
    #     print(f"{x},{y}")
    # print("\n")

    return x_fit_circle, y_fit_circle, fitting_error, R, xc, yc

def calculate_curvature(f, x):
    """
    计算给定样条曲线在指定 x 值处的曲率。

    参数：
    f (callable): 函数或插值对象，接受 x 数组作为参数并返回对应的因变量数组。
    x (numpy.ndarray): 自变量数组。

    返回：
    numpy.ndarray: 对应于 x 数组的曲率数组。

    注意：
    - 计算曲率公式使用一阶和二阶导数。
    """
    # 计算一阶和二阶导数
    f_prime = f.derivative()
    f_second = f_prime.derivative()
    y_prime_segment = f_prime(x)
    y_second_segment = f_second(x)
    
    # 计算曲率
    curvature = np.abs(y_second_segment) / (1 + y_prime_segment**2)**(3/2)
    
    return curvature

# find_zero_second_derivatives函数用于在给定的数据点 (x, y_dedrivative) 中查找二阶导数为零的段落和二阶导数不为零的段落。

def find_zero_second_derivatives(x, y_dedrivative, tolerance):
    """
    在数据点 (x, y_dedrivative) 中查找二阶导数为零的段落和二阶导数不为零的段落。

    Parameters:
    x : numpy.ndarray
        输入数据的 x 值数组。
    y_dedrivative : numpy.ndarray
        输入数据的二阶导数值数组。
    tolerance : float
        允许的二阶导数接近零的容差。

    Returns:
    tuple
        包含二阶导数为零段落和二阶导数不为零段落的元组。
        每个段落表示为 (start_index, end_index) 形式的元组。
    """

    # 找到二阶导数为零的段落
    zero_crossings = np.where(np.abs(y_dedrivative) < tolerance)[0]  # 找到二阶导数绝对值小于指定容差的索引
    zero_segments1 = np.split(zero_crossings, np.where(np.diff(zero_crossings) != 1)[0] + 1)  # 将连续的索引分割成段落
    zero_segments1 = [(segment[0], segment[-1]) for segment in zero_segments1 if len(segment) > 0]  # 记录每个段落的起始和结束索引
    # 打印每个段落的详细信息
    print("零", zero_segments1)
    zero_segments = [segment for segment in zero_segments1 if segment[0] != segment[1] and abs(segment[0] - segment[1]) > 5]
    #zero_segments = zero_segments1
    print("零1", zero_segments)



    # 找到二阶导数不为零的段落
    nonzero_segments = []
    start_index = 0
    for segment in zero_segments:
        end_index = segment[0]
        if start_index < end_index:
            nonzero_segments.append((start_index, end_index))  # 添加二阶导数不为零的段落的起始和结束索引
        start_index = segment[-1]
    if start_index < len(x):
        nonzero_segments.append((start_index, len(x)))  # 添加最后一个段落的起始和结束索引

    segments_to_remove_index = []  # 记录需要删除的段落索引
    

    # 处理重叠段落
    for i, nonzero_segment in enumerate(nonzero_segments):
        if nonzero_segment[1] - nonzero_segment[0] > 2:
            continue
        # 找到尾数等于nonzero_segment[0]的段落索引，将其尾数改为nonzero_segment[1]
        segment_index = next((i for i, segment in enumerate(zero_segments) if segment[1] == nonzero_segment[0]), None)
        if segment_index is not None:
            zero_segments[segment_index] = (zero_segments[segment_index][0], nonzero_segment[1])  # 更新重叠段落的结束索引
            segments_to_remove_index.append(i)  # 记录需要删除的段落索引

    # 删除需要删除的段落
    for index in sorted(segments_to_remove_index, reverse=True):
        del nonzero_segments[index]  # 删除二阶导数不为零的段落中的重叠部分

    return zero_segments, nonzero_segments  # 返回二阶导数为零和二阶导数不为零的段落列表

def find_zero_derivatives(x, y_smooth, y_prime, y_second, y_prime_diff, tolerance):
    """
    在数据点 (x, y_dedrivative) 中查找二阶导数为零的段落和二阶导数不为零的段落。

    Parameters:
    x : numpy.ndarray
        输入数据的 x 值数组。
    y_dedrivative : numpy.ndarray
        输入数据的二阶导数值数组。
    tolerance : float
        允许的二阶导数接近零的容差。

    Returns:
    tuple
        包含二阶导数为零段落和二阶导数不为零段落的元组。
        每个段落表示为 (start_index, end_index) 形式的元组。
    """
    
    # 找到二阶导数为零的段落
    zero_crossings = np.where(np.abs(y_prime_diff) < tolerance)[0]  # 找到二阶导数绝对值小于指定容差的索引
    zero_segments1 = np.split(zero_crossings, np.where(np.diff(zero_crossings) != 1)[0] + 1)  # 将连续的索引分割成段落
    zero_segments1 = [(segment[0], segment[-1]) for segment in zero_segments1 if len(segment) > 0]  # 记录每个段落的起始和结束索引
    # 打印每个段落的详细信息
    print("零", zero_segments1)
    zero_segments = [segment for segment in zero_segments1 if segment[0] != segment[1] and abs(segment[0] - segment[1]) > 5]
    #zero_segments = zero_segments1
    print("零1", zero_segments)

    # 找到二阶导数不为零的段落
    nonzero_segments = []
    start_index = 0
    for segment in zero_segments:
        end_index = segment[0]
        if start_index < end_index:
            nonzero_segments.append((start_index, end_index))  # 添加二阶导数不为零的段落的起始和结束索引
        start_index = segment[-1]
    if start_index < len(x):
        nonzero_segments.append((start_index, len(x)))  # 添加最后一个段落的起始和结束索引

    segments_to_remove_index = []  # 记录需要删除的段落索引
    
    # 处理重叠段落
    for i, nonzero_segment in enumerate(nonzero_segments):
        if nonzero_segment[1] - nonzero_segment[0] > 2:
            continue
        # 找到尾数等于nonzero_segment[0]的段落索引，将其尾数改为nonzero_segment[1]
        segment_index = next((i for i, segment in enumerate(zero_segments) if segment[1] == nonzero_segment[0]), None)
        if segment_index is not None:
            zero_segments[segment_index] = (zero_segments[segment_index][0], nonzero_segment[1])  # 更新重叠段落的结束索引
            segments_to_remove_index.append(i)  # 记录需要删除的段落索引

    # 删除需要删除的段落
    for index in sorted(segments_to_remove_index, reverse=True):
        del nonzero_segments[index]  # 删除二阶导数不为零的段落中的重叠部分

    return zero_segments, nonzero_segments  # 返回二阶导数为零和二阶导数不为零的段落列表

def compute_derivatives(f, x):
    """
    计算函数 f 在给定点 x 处的一阶、二阶和三阶导数值。

    Parameters:
    f : callable
        可调用的函数或插值对象，可以计算导数。
    x : numpy.ndarray
        输入数据的 x 值数组。

    Returns:
    tuple
        包含一阶、二阶和三阶导数值的元组。
    """
    f_prime = f.derivative()  # 计算一阶导数
    f_second = f_prime.derivative()  # 计算二阶导数
    f_third = f_second.derivative()  # 计算三阶导数
    y_prime = f_prime(x)  # 计算一阶导数值
    y_second = f_second(x)  # 计算二阶导数值
    y_third = f_third(x)  # 计算三阶导数值  
    return y_prime, y_second, y_third  # 返回一阶、二阶和三阶导数值的元组

def extract_data_from_excel(file_path):
    """
    从指定的 Excel 文件中提取数据。

    Parameters:
    file_path : str
        Excel 文件的路径。

    Returns:
    numpy.ndarray
        包含从 Excel 中提取的数据的 NumPy 数组，每行包括"桩号"、"桩号数字"、"X 值"和"Y 值"。
    """
    # 将相对路径转换为绝对路径
    abs_file_path = os.path.abspath(file_path)
    
    # 加载 Excel 文件
    wb = load_workbook(abs_file_path)
    ws = wb.active
    # 用于存储结果的列表
    data = []

    # 遍历工作表中的每一行
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            # 检查单元格的值是否包含"k"
            if isinstance(cell.value, str) and "k" in cell.value.lower():
                # 提取当前单元格及其右侧两个单元格的值
                stake = cell.value
                stake_number = extract_number(stake)  # 匹配整个单词
    
                x_value = ws.cell(row=cell.row, column=cell.column + 1).value
                y_value = ws.cell(row=cell.row, column=cell.column + 2).value
                if x_value is not None and y_value is not None:
                    # 将数据添加到结果列表中
                    data.append([stake, stake_number, float(x_value), float(y_value)])
    
    # 关闭 Excel 文件
    wb.close()

    # 根据 stake_number 进行排序
    sorted_data = sorted(data, key=lambda x: x[1])
    
    return np.asarray(sorted_data)

def extract_number(text):
    """
    从给定的文本中提取并组合数字。

    参数:
    text (str): 包含数字的字符串。

    返回:
    float: 提取并组合后的浮点数。如果未找到匹配的数字，则返回 None。
    """
    # 使用正则表达式匹配两个部分的数字（例如 "123+45.67"）
    number_match = re.search(r'(\d+)\+(\d+\.?\d*)', text)
    
    if number_match:
        # 如果找到匹配的数字，提取第一部分和第二部分
        number_part1 = number_match.group(1)
        number_part2 = number_match.group(2)
        
        # 将两个部分组合成一个字符串
        combined_number = number_part1 + number_part2
        
        # 将组合后的字符串转换为浮点数并返回
        return float(combined_number)
    else:
        # 如果未找到匹配的数字，返回 None
        return None
    
def extract_data_from_excel_lon(file_path):
    """
    从指定的 Excel 或 CSV 文件中提取数据。

    Parameters:
    file_path : str
        文件的路径，可以是 Excel 或 CSV 文件。

    Returns:
    numpy.ndarray
        包含从文件中提取的数据的 NumPy 数组，每行包括"桩号"和"高程值"。
    """
    # 将相对路径转换为绝对路径
    abs_file_path = os.path.abspath(file_path)
    
    # 确定文件类型
    file_extension = os.path.splitext(abs_file_path)[1].lower()
    
    # 用于存储结果的列表
    data = []

    if file_extension in ['.xls', '.xlsx']:
        from openpyxl import load_workbook
        # 加载 Excel 文件
        wb = load_workbook(abs_file_path)
        ws = wb.active

        # 遍历工作表中的每一行
        for row in ws.iter_rows(min_row=1, max_col=2, max_row=ws.max_row):
            stake = row[0].value
            elevation = row[1].value

            if isinstance(stake, str) and isinstance(elevation, (int, float)):
                stake_number = extract_number(stake)
                if stake_number is not None:
                    data.append([stake, float(stake_number), float(elevation)])
        
        # 关闭 Excel 文件
        wb.close()

    elif file_extension == '.csv':
        # 加载 CSV 文件
        df = pd.read_csv(abs_file_path)

        # 假设 CSV 文件的前两列分别是桩号和高程值
        for index, row in df.iterrows():
            stake = row[0]
            elevation = row[1]

            if isinstance(stake, str) and isinstance(elevation, (int, float)):
                stake_number = extract_number(stake)
                if stake_number is not None:
                    data.append([stake, float(stake_number), float(elevation)])

    # 根据 stake_number 进行排序
    sorted_data = sorted(data, key=lambda x: x[1])
    
    return np.asarray(sorted_data)
    
def clothoid_A(x1, y1, theta1, x2, y2, theta2):
    """
    计算连接两点并给定切向方向的缓和曲线（欧拉螺线）的参数A。
    参数:
    x1, y1: float
        缓和曲线起点的坐标。
    theta1: float
        起点的切向方向（弧度）。
    x2, y2: float
        缓和曲线终点的坐标。
    theta2: float
        终点的切向方向（弧度）。

    返回值:
    A_optimal: float
        缓和曲线的最优参数A。
    """
    def objective(L):
        """
        目标函数，用于最小化计算的缓和曲线终点和方向与期望的终点和方向之间的误差。
        minimize 使用的方法是 BFGS（Broyden-Fletcher-Goldfarb-Shanno 算法），这是一种准牛顿法，用于无约束优化问题

        参数:
        L: float
            缓和曲线的长度。

        返回值:
        error: float
            欧几里得距离误差和角度误差之和。
        """
        A = np.sqrt(L)  # 根据长度 L 计算参数 A
        # 计算缓和曲线终点的 x 坐标
        x_clothoid = x1 + (L**2 / (2 * A**2)) * np.cos(theta1)
        # 计算缓和曲线终点的 y 坐标
        y_clothoid = y1 + (L**2 / (2 * A**2)) * np.sin(theta1)
        # 计算缓和曲线终点的切向方向
        theta_clothoid = theta1 + L / A**2
        # 计算缓和曲线终点与期望终点之间的误差
        error = np.sqrt((x_clothoid - x2)**2 + (y_clothoid - y2)**2) + np.abs(theta_clothoid - theta2)
        return error

    # 使用最小化函数找到最优的 L
    result = minimize(objective, 1.0, bounds=[(0, None)])
    # 提取最优 L 的值
    L_optimal = result.x[0]
    # 根据最优 L 计算最优参数 A
    A_optimal = np.sqrt(L_optimal)
    return A_optimal

def calculate_length(points):
    """
    计算点集中的点依次相连的总长度

    参数:
    points (numpy.ndarray): 形状为 (n, 2) 的二维数组，其中 n 是点的数量，数组的每一行表示一个点的 (x, y) 坐标

    返回:
    float: 点集中的点依次相连的总长度
    """
    total_length = 0
    for i in range(1, len(points)):
        total_length += np.sqrt((points[i, 0] - points[i-1, 0])**2 + (points[i, 1] - points[i-1, 1])**2)
    return total_length

def calculate_A(L, R):
    # 计算缓和曲线终点的曲率
    curvature_end = 1 / R
    
    # 计算螺距 A
    A = np.sqrt(L / curvature_end)
    
    return A